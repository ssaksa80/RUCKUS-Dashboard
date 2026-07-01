"""Daily Excel report — KPI overview + per-domain sheets with charts.

``build_report`` accepts either a ``ReportModel`` (reports/model.py) — the
generic, all-module path — or the legacy normalized dict
``{"aps": [...], "clients": [...], "alarms": [...], "switches": [...]}``.
Both render the curated chart sheets plus a model-driven Coverage sheet and one
generic sheet per module. Returns xlsx bytes (openpyxl)."""
from __future__ import annotations

import io
import re
import time
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .model import ModuleReport, ReportModel

_HEAD = Font(bold=True, color="FFFFFF")
_HEAD_FILL = PatternFill("solid", fgColor="22A6B3")

_ILLEGAL_SHEET = re.compile(r"[:\\/?*\[\]]")


def _header(ws, row: int, labels: list[str]) -> None:
    for col, label in enumerate(labels, start=1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL


def _autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _safe_sheet_name(title: str, used: set[str]) -> str:
    """Excel sheet names: <=31 chars, none of :\\/?*[]; unique within a book."""
    base = _ILLEGAL_SHEET.sub(" ", str(title or "Sheet")).strip()[:31] or "Sheet"
    name = base
    n = 2
    while name in used:
        suffix = f" ({n})"
        name = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(name)
    return name


def _wrap_legacy(data: dict[str, Any]) -> ReportModel:
    """Adapt the legacy {aps,clients,alarms,switches} dict to a minimal model
    so the model-driven Overview/Coverage render alongside the curated sheets."""
    mods: list[ModuleReport] = []
    titles = {"aps": ("Access Points", "Wireless"),
              "clients": ("Clients", "Wireless"),
              "alarms": ("Alarms", "Wireless"),
              "switches": ("Switches", "Switching")}
    for slug, (title, group) in titles.items():
        rows = list(data.get(slug) or [])
        mods.append(ModuleReport(slug=slug, title=title, group=group,
                                 status="ok", rows=rows, row_total=len(rows)))
    return ReportModel(generated_at=time.strftime("%Y-%m-%dT%H:%M UTC",
                                                  time.gmtime()),
                       connection_label="", modules=mods)


def _legacy_from_model(model: ReportModel) -> dict[str, Any]:
    """Pull the four curated domains' rows out of a model for the chart sheets."""
    out: dict[str, Any] = {}
    for slug in ("aps", "clients", "alarms", "switches"):
        rep = model.by_slug(slug)
        out[slug] = list(rep.rows) if rep else []
    return out


def build_report(data_or_model) -> bytes:
    """Render xlsx bytes from a ReportModel (new) or the legacy
    {aps,clients,...} dict (curated sheets + model-driven Overview/Coverage)."""
    if isinstance(data_or_model, ReportModel):
        model = data_or_model
        legacy = _legacy_from_model(model)
    else:
        legacy = data_or_model or {}
        model = _wrap_legacy(legacy)
    wb = Workbook()
    _build_curated(wb, legacy)
    _build_coverage(wb, model)
    _build_module_sheets(wb, model)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_curated(wb: Workbook, data: dict[str, Any]) -> None:
    aps = data.get("aps") or []
    clients = data.get("clients") or []
    alarms = data.get("alarms") or []
    switches = data.get("switches") or []

    # ── Overview ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Overview"
    ws["A1"] = "RUCKUS DSO Daily Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = time.strftime("%Y-%m-%d %H:%M UTC", time.gmtime())
    aps_off = sum(1 for a in aps if a.get("status") == "offline")
    sw_off = sum(1 for s in switches if str(s.get("status")).lower() not in
                 ("online", "in_service"))
    crit = sum(int(a.get("count") or 0) for a in alarms
               if a.get("severity") == "critical")
    rows = [
        ("Access points (total)", len(aps)),
        ("Access points offline", aps_off),
        ("Wireless clients", len(clients)),
        ("Clients with poor signal", sum(1 for c in clients if c.get("quality") == "poor")),
        ("Switches (total)", len(switches)),
        ("Switches offline", sw_off),
        ("Active alarms", sum(int(a.get("count") or 0) for a in alarms)),
        ("Critical alarms", crit),
    ]
    _header(ws, 4, ["Metric", "Value"])
    for i, (k, v) in enumerate(rows, start=5):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    _autofit(ws, [34, 14])

    # ── APs by Zone (+ bar chart) ─────────────────────────────────────────
    ws = wb.create_sheet("APs by Zone")
    by_zone: dict[str, dict[str, int]] = {}
    for a in aps:
        z = str(a.get("zone") or "Unknown")
        d = by_zone.setdefault(z, {"total": 0, "offline": 0})
        d["total"] += 1
        if a.get("status") == "offline":
            d["offline"] += 1
    _header(ws, 1, ["Zone", "APs", "Offline"])
    for i, (z, d) in enumerate(sorted(by_zone.items()), start=2):
        ws.cell(row=i, column=1, value=z)
        ws.cell(row=i, column=2, value=d["total"])
        ws.cell(row=i, column=3, value=d["offline"])
    _autofit(ws, [30, 10, 10])
    if by_zone:
        chart = BarChart()
        chart.title = "APs per zone (total vs offline)"
        chart.height, chart.width = 10, 24
        last = 1 + len(by_zone)
        chart.add_data(Reference(ws, min_col=2, max_col=3, min_row=1, max_row=last),
                       titles_from_data=True)
        chart.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(chart, "E2")

    # ── Clients (+ pie chart by band, top talkers) ───────────────────────
    ws = wb.create_sheet("Clients")
    bands: dict[str, int] = {}
    for c in clients:
        bands[str(c.get("band") or "—")] = bands.get(str(c.get("band") or "—"), 0) + 1
    _header(ws, 1, ["Band", "Clients"])
    for i, (b, n) in enumerate(sorted(bands.items()), start=2):
        ws.cell(row=i, column=1, value=b)
        ws.cell(row=i, column=2, value=n)
    if bands:
        pie = PieChart()
        pie.title = "Clients by band"
        pie.height, pie.width = 9, 12
        last = 1 + len(bands)
        pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=last),
                     titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(pie, "D2")
    top = sorted(clients, key=lambda c: int(c.get("rx_bytes") or 0) +
                 int(c.get("tx_bytes") or 0), reverse=True)[:10]
    base = 4 + len(bands)
    ws.cell(row=base - 1, column=1, value="Top talkers").font = Font(bold=True)
    _header(ws, base, ["Host", "MAC", "SSID", "AP", "RX bytes", "TX bytes"])
    for i, c in enumerate(top, start=base + 1):
        for col, key in enumerate(("hostname", "mac", "ssid", "ap",
                                   "rx_bytes", "tx_bytes"), start=1):
            ws.cell(row=i, column=col, value=c.get(key))
    _autofit(ws, [22, 20, 16, 20, 14, 14])

    # ── Alarms (+ pie chart) ─────────────────────────────────────────────
    ws = wb.create_sheet("Alarms")
    sev: dict[str, int] = {}
    for a in alarms:
        s = str(a.get("severity") or "unknown")
        sev[s] = sev.get(s, 0) + int(a.get("count") or 1)
    _header(ws, 1, ["Severity", "Count"])
    for i, (s, n) in enumerate(sorted(sev.items()), start=2):
        ws.cell(row=i, column=1, value=s)
        ws.cell(row=i, column=2, value=n)
    if sev:
        pie = PieChart()
        pie.title = "Alarms by severity"
        pie.height, pie.width = 9, 12
        last = 1 + len(sev)
        pie.add_data(Reference(ws, min_col=2, min_row=1, max_row=last),
                     titles_from_data=True)
        pie.set_categories(Reference(ws, min_col=1, min_row=2, max_row=last))
        ws.add_chart(pie, "D2")
    base = 4 + len(sev)
    ws.cell(row=base - 1, column=1, value="Active alarms").font = Font(bold=True)
    _header(ws, base, ["Severity", "Category", "Source", "Message", "Count"])
    for i, a in enumerate(alarms[:50], start=base + 1):
        for col, key in enumerate(("severity", "category", "source",
                                   "message", "count"), start=1):
            ws.cell(row=i, column=col, value=a.get(key))
    _autofit(ws, [12, 14, 24, 48, 8])

    # ── Switches (+ traffic bar chart) ───────────────────────────────────
    ws = wb.create_sheet("Switches")
    _header(ws, 1, ["Switch", "IP", "Model", "Firmware", "Status",
                    "Ports up", "Ports total"])
    for i, s in enumerate(switches, start=2):
        for col, key in enumerate(("name", "ip", "model", "fw", "status",
                                   "ports_online", "ports_total"), start=1):
            ws.cell(row=i, column=col, value=s.get(key))
    _autofit(ws, [24, 16, 16, 18, 10, 10, 10])

    # ── Offline devices ──────────────────────────────────────────────────
    ws = wb.create_sheet("Offline Devices")
    _header(ws, 1, ["Type", "Name", "Where", "Identifier"])
    r = 2
    for a in aps:
        if a.get("status") == "offline":
            ws.cell(row=r, column=1, value="AP")
            ws.cell(row=r, column=2, value=a.get("name"))
            ws.cell(row=r, column=3, value=a.get("zone"))
            ws.cell(row=r, column=4, value=a.get("mac"))
            r += 1
    for s in switches:
        if str(s.get("status")).lower() not in ("online", "in_service"):
            ws.cell(row=r, column=1, value="Switch")
            ws.cell(row=r, column=2, value=s.get("name"))
            ws.cell(row=r, column=3, value=s.get("group"))
            ws.cell(row=r, column=4, value=s.get("mac") or s.get("id"))
            r += 1
    _autofit(ws, [10, 26, 24, 22])


def _build_coverage(wb: Workbook, model: ReportModel) -> None:
    ws = wb.create_sheet("Coverage")
    ws["A1"] = "Module coverage"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated {model.generated_at}"
    _header(ws, 4, ["Module", "Group", "Status", "Rows", "Errors", "Note"])
    for i, m in enumerate(model.modules, start=5):
        ws.cell(row=i, column=1, value=m.title)
        ws.cell(row=i, column=2, value=m.group)
        ws.cell(row=i, column=3, value=m.status)
        ws.cell(row=i, column=4, value=m.row_total)
        ws.cell(row=i, column=5, value=len(m.errors))
        ws.cell(row=i, column=6, value=m.note or "")
    _autofit(ws, [26, 14, 10, 8, 8, 40])


_LIST_ROW_CAP = 1000


def _fmt_value(value, kind: str):
    """Light, render-only formatting matching the dashboard idioms."""
    if value is None:
        return ""
    if kind == "bytes":
        try:
            return _human_bytes_xl(int(value))
        except (TypeError, ValueError):
            return value
    return value


def _human_bytes_xl(n: int) -> str:
    v = float(n or 0)
    if v <= 0:
        return "0 B"
    for unit in ("B", "KB", "MB", "GB", "TB", "PB"):
        if v < 1024:
            return f"{v:.0f} {unit}" if unit == "B" else f"{v:.1f} {unit}"
        v /= 1024
    return f"{v:.1f} EB"


def _kv_block(ws, row: int, title: str, mapping: dict) -> int:
    """Write a 'title' header then key/value rows. Returns the next free row."""
    ws.cell(row=row, column=1, value=title).font = Font(bold=True)
    row += 1
    for k, v in (mapping or {}).items():
        ws.cell(row=row, column=1, value=str(k))
        ws.cell(row=row, column=2, value=v if not isinstance(v, (dict, list))
                else str(v))
        row += 1
    return row + 1


def _build_module_sheets(wb: Workbook, model: ReportModel) -> None:
    used = {ws_title for ws_title in wb.sheetnames}
    for m in model.modules:
        ws = wb.create_sheet(_safe_sheet_name(m.title, used))
        ws["A1"] = m.title
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Status: {m.status}"
        if m.note:
            ws["A3"] = m.note
        if m.filters_applied:
            applied = ", ".join(f"{k}={v}" for k, v in m.filters_applied.items()
                                if v)
            ws["B2"] = f"Filters: {applied}" if applied else "Filters: none"
        row = 5
        row = _kv_block(ws, row, "Summary", m.summary)

        # List table.
        ws.cell(row=row, column=1, value="List").font = Font(bold=True)
        row += 1
        if m.columns and m.rows:
            labels = [c.label for c in m.columns]
            _header(ws, row, labels)
            row += 1
            shown = m.rows[:_LIST_ROW_CAP]
            for r in shown:
                for col, c in enumerate(m.columns, start=1):
                    ws.cell(row=row, column=col,
                            value=_fmt_value(r.get(c.key), c.kind))
                row += 1
            extra = len(m.rows) - len(shown)
            if extra > 0:
                ws.cell(row=row, column=1, value=f"+{extra} more rows (capped)")
                row += 1
        else:
            ws.cell(row=row, column=1,
                    value="(no list)" if not m.rows else "(no columns declared)")
            row += 1
        row += 1

        # Raw field-map samples.
        if m.raw_samples:
            ws.cell(row=row, column=1, value="Raw field map").font = Font(bold=True)
            row += 1
            for i, sample in enumerate(m.raw_samples, start=1):
                row = _kv_block(ws, row, f"Sample {i}", sample)

        # Drill samples.
        if m.drill_samples:
            ws.cell(row=row, column=1, value="Drill samples").font = Font(bold=True)
            row += 1
            for d in m.drill_samples:
                if d.error:
                    ws.cell(row=row, column=1,
                            value=f"{d.entity_id}: error — {d.error}")
                    row += 2
                    continue
                for section, payload in (d.sections or {}).items():
                    flat = payload if isinstance(payload, dict) else {"value": payload}
                    row = _kv_block(ws, row, f"{d.entity_id} · {section}", flat)
        _autofit(ws, [28, 32])
