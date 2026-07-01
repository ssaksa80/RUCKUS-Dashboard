from ruckus_dashboard.app import create_app


def _app(tmp_path):
    app = create_app({"SECRET_KEY": "t", "RUCKUS_ENABLE_NEW_UI": True})
    app.instance_path = str(tmp_path)
    return app


def _login(c):
    c.get("/")
    with c.session_transaction() as s:
        s["auth"] = True
        s["connection_ids"] = []
        return s["csrf_token"]


def test_notifications_api_requires_auth(tmp_path):
    app = _app(tmp_path)
    with app.test_client() as c:
        assert c.get("/api/notifications/config").status_code == 401
        assert c.post("/api/notifications/config", json={}).status_code == 401
        assert c.post("/api/notifications/test").status_code == 401
        assert c.get("/api/reports/generate").status_code == 401


def test_notifications_config_roundtrip_masks_password(tmp_path):
    app = _app(tmp_path)
    with app.test_client() as c:
        csrf = _login(c)
        r = c.post("/api/notifications/config",
                   json={"smtp": {"host": "mail.x", "password": "hunter2"},
                         "alerts": {"enabled": True, "recipients": ["a@x"]}},
                   headers={"X-CSRF-Token": csrf})
        assert r.status_code == 200
        body = r.get_json()
        assert body["smtp"]["password"] == "********"
        assert "hunter2" not in r.get_data(as_text=True)
        got = c.get("/api/notifications/config").get_json()
        assert got["smtp"]["host"] == "mail.x"
        assert got["alerts"]["recipients"] == ["a@x"]


def test_test_email_route_uses_mailer(tmp_path, monkeypatch):
    import ruckus_dashboard.routes.notifications as notif_routes
    calls = {}

    def fake_send(cfg, pw, recipients, subject, body, **kw):
        calls["recipients"] = recipients
        calls["subject"] = subject

    monkeypatch.setattr(notif_routes, "send_email", fake_send)
    app = _app(tmp_path)
    with app.test_client() as c:
        csrf = _login(c)
        c.post("/api/notifications/config",
               json={"smtp": {"host": "mail.x"},
                     "alerts": {"recipients": ["noc@x"]}},
               headers={"X-CSRF-Token": csrf})
        r = c.post("/api/notifications/test", headers={"X-CSRF-Token": csrf})
        assert r.status_code == 200
        assert r.get_json()["sent"] is True
        assert calls["recipients"] == ["noc@x"]


def test_notifications_page_renders(tmp_path):
    app = _app(tmp_path)
    with app.test_client() as c:
        _login(c)
        r = c.get("/notifications")
        assert r.status_code == 200
        assert b"data-notifications" in r.data
        assert b"notifications.js" in r.data


def test_notifications_config_roundtrip_includes_sp2_fields(tmp_path):
    """Config GET/POST roundtrip carries new SP2 alert fields."""
    app = _app(tmp_path)
    with app.test_client() as c:
        csrf = _login(c)
        r = c.post("/api/notifications/config",
                   json={"smtp": {"host": "mail.x", "password": "hunter2"},
                         "alerts": {
                             "enabled": True,
                             "recipients": ["a@x"],
                             "recovery": False,
                             "debounce_seconds": 60,
                             "group_by": "none",
                         }},
                   headers={"X-CSRF-Token": csrf})
        assert r.status_code == 200
        body = r.get_json()
        alerts = body["alerts"]
        assert alerts["recovery"] is False
        assert alerts["debounce_seconds"] == 60
        assert alerts["group_by"] == "none"
        # Password still masked.
        assert body["smtp"]["password"] == "********"

        # GET returns same values.
        got = c.get("/api/notifications/config").get_json()
        assert got["alerts"]["recovery"] is False
        assert got["alerts"]["debounce_seconds"] == 60


def test_test_alert_email_sends_grouped_body(tmp_path, monkeypatch):
    """kind='alerts' test email sends subject+body using the configured recipients."""
    import ruckus_dashboard.routes.notifications as notif_routes
    calls = {}

    def fake_send(cfg, pw, recipients, subject, body, **kw):
        calls["recipients"] = recipients
        calls["subject"] = subject
        calls["body"] = body

    monkeypatch.setattr(notif_routes, "send_email", fake_send)
    app = _app(tmp_path)
    with app.test_client() as c:
        csrf = _login(c)
        c.post("/api/notifications/config",
               json={"smtp": {"host": "mail.x"},
                     "alerts": {"recipients": ["noc@x"]}},
               headers={"X-CSRF-Token": csrf})
        r = c.post("/api/notifications/test",
                   json={"kind": "alerts"},
                   headers={"X-CSRF-Token": csrf})
        assert r.status_code == 200
        assert r.get_json()["sent"] is True
        assert calls["recipients"] == ["noc@x"]
        # Test alert body mentions the outage channel (not just "smtp works").
        assert "alert" in calls["body"].lower() or "RUCKUS DSO" in calls["subject"]


def _authed_with_conn(tmp_path):
    """App + one stored SmartZone connection; returns (app, csrf)."""
    from ruckus_dashboard.auth.session_store import ConnectionConfig
    app = _app(tmp_path)
    conn = ConnectionConfig(platform="smartzone", api_base="https://sz/wsg/api/public",
                            display_name="SZ-LAB", auth_token="t",
                            api_version="v11_0", verify_tls=False,
                            token_expires_at=9999999999)
    cid = app.connection_store.put(conn)
    # Production path: capability ops live in the per-connection registry keyed
    # by the session's connection id — NOT a process-global ``app.available_ops``
    # (that attribute was removed in SP7; the report routes read the registry).
    app.capability_registry.set_for(
        cid, {("POST", "/query/client"), ("POST", "/query/ap")})
    with app.test_client() as c:
        c.get("/")
        with c.session_transaction() as s:
            s["auth"] = True
            s["connection_ids"] = [cid]
            csrf = s["csrf_token"]
        yield c, csrf


def test_reports_tab_requires_auth(tmp_path):
    app = _app(tmp_path)
    with app.test_client() as c:
        assert c.post("/api/reports/tab", json={"slug": "clients"}).status_code == 401


def test_reports_tab_requires_csrf(tmp_path):
    for c, _csrf in [next(_authed_with_conn(tmp_path))]:
        r = c.post("/api/reports/tab", json={"slug": "clients"})
        assert r.status_code == 400          # missing X-CSRF-Token


def test_reports_tab_unknown_slug_404(tmp_path):
    for c, csrf in [next(_authed_with_conn(tmp_path))]:
        r = c.post("/api/reports/tab", json={"slug": "nope"},
                   headers={"X-CSRF-Token": csrf})
        assert r.status_code == 404


def test_reports_tab_happy_path_emails_one_module(tmp_path, monkeypatch):
    import ruckus_dashboard.routes.notifications as notif_routes
    import ruckus_dashboard.reports.collect as collect_mod
    calls = {}

    def fake_send(cfg, pw, recipients, subject, body, **kw):
        calls["recipients"] = recipients
        calls["subject"] = subject
        calls["filename"] = kw.get("filename")
        calls["has_attachment"] = kw.get("attachment") is not None

    captured = {}
    real_collect = collect_mod.collect_report_model

    def spy_collect(*a, **kw):
        captured["slugs"] = kw.get("slugs")
        captured["filters_by_slug"] = kw.get("filters_by_slug")
        return real_collect(*a, **kw)

    monkeypatch.setattr(notif_routes, "send_email", fake_send)
    monkeypatch.setattr(notif_routes, "collect_report_model", spy_collect)

    gen = _authed_with_conn(tmp_path)
    c, csrf = next(gen)
    # Configure report recipients.
    c.post("/api/notifications/config",
           json={"smtp": {"host": "mail.x"}, "report": {"recipients": ["noc@x"]}},
           headers={"X-CSRF-Token": csrf})
    # Stub the clients fetcher so no HTTP happens.
    import ruckus_dashboard.modules as modmod
    import dataclasses
    original = modmod.MODULES["clients"]
    modmod.MODULES["clients"] = dataclasses.replace(
        original,
        fetcher=lambda ctx: {"items": [{"id": "a", "band": "5 GHz"},
                                       {"id": "b", "band": "2.4 GHz"}]},
        drill_fetcher=None)
    try:
        r = c.post("/api/reports/tab",
                   json={"slug": "clients", "filters": {"band": "5 GHz"}},
                   headers={"X-CSRF-Token": csrf})
        assert r.status_code == 200, r.get_data(as_text=True)
        body = r.get_json()
        assert body["sent"] is True
        assert body["slug"] == "clients"
        assert calls["recipients"] == ["noc@x"]
        assert "clients" in calls["filename"]
        assert calls["has_attachment"] is True
        # Filters forwarded into the collector for that slug only.
        assert captured["slugs"] == ("clients",)
        assert captured["filters_by_slug"] == {"clients": {"band": "5 GHz"}}
    finally:
        modmod.MODULES["clients"] = original


def test_reports_tab_gated_module_granted_via_registry_not_422(tmp_path, monkeypatch):
    """Production-path guard: the capability gate on ``/api/reports/tab`` reads
    the per-connection ``capability_registry`` (seeded for the session's cid),
    NOT the removed process-global ``app.available_ops``.

    'clients' requires ("POST","/query/client"), which the fixture grants in the
    registry. Before the SP7 wiring fix the route read ``getattr(app,
    "available_ops", set())`` — always empty in production — and returned 422
    even though the tab renders in the UI. After the fix the gate is satisfied
    and the request proceeds (email send is stubbed so we assert the 200/sent).
    """
    import ruckus_dashboard.routes.notifications as notif_routes
    import ruckus_dashboard.modules as modmod
    import dataclasses

    monkeypatch.setattr(notif_routes, "send_email",
                        lambda *a, **kw: None)

    gen = _authed_with_conn(tmp_path)
    c, csrf = next(gen)
    c.post("/api/notifications/config",
           json={"smtp": {"host": "mail.x"}, "report": {"recipients": ["noc@x"]}},
           headers={"X-CSRF-Token": csrf})
    original = modmod.MODULES["clients"]
    modmod.MODULES["clients"] = dataclasses.replace(
        original, fetcher=lambda ctx: {"items": [{"id": "a"}]}, drill_fetcher=None)
    try:
        r = c.post("/api/reports/tab", json={"slug": "clients"},
                   headers={"X-CSRF-Token": csrf})
        # The load-bearing assertion: the capability gate did NOT reject a module
        # the connection actually has (this is exactly the 422 the operator saw).
        assert r.status_code != 422, r.get_data(as_text=True)
        assert r.status_code == 200, r.get_data(as_text=True)
        assert r.get_json()["sent"] is True
    finally:
        modmod.MODULES["clients"] = original


def test_reports_tab_disabled_module_returns_422(tmp_path):
    gen = _authed_with_conn(tmp_path)
    c, csrf = next(gen)
    # 'rogues' requires ("POST","/query/roguesInfoList"), which the fixture does
    # NOT grant in the registry — a module whose caps are not granted still 422s.
    r = c.post("/api/reports/tab", json={"slug": "rogues"},
               headers={"X-CSRF-Token": csrf})
    assert r.status_code == 422
    assert r.get_json()["sent"] is False


def test_reports_generate_covers_all_modules_no_crash(tmp_path, monkeypatch):
    """/api/reports/generate runs the 19-module collector; topology/overview
    shapes must not crash the workbook (regression for the 4-module blind spot)."""
    import io
    import openpyxl
    import ruckus_dashboard.modules as modmod
    import dataclasses

    gen = _authed_with_conn(tmp_path)
    c, _csrf = next(gen)
    # Make every module enabled + cheap; keep topology/overview real shapes.
    originals = dict(modmod.MODULES)
    try:
        for slug, spec in list(modmod.MODULES.items()):
            if slug in ("topology", "overview"):
                modmod.MODULES[slug] = dataclasses.replace(
                    spec, requires_capabilities=())
                continue
            modmod.MODULES[slug] = dataclasses.replace(
                spec,
                fetcher=lambda ctx, s=slug: {"items": [{"id": f"{s}-1"}],
                                             "raw_count": 1},
                drill_fetcher=None, requires_capabilities=())
        # topology fetcher returns its graph shape; stub to avoid HTTP.
        modmod.MODULES["topology"] = dataclasses.replace(
            modmod.MODULES["topology"],
            fetcher=lambda ctx: {"nodes": [{"id": "controller"}], "edges": [],
                                 "items": []})
        r = c.get("/api/reports/generate")
        assert r.status_code == 200
        wb = openpyxl.load_workbook(io.BytesIO(r.data))
        assert "Coverage" in wb.sheetnames
        # Every module title shows up on the Coverage sheet.
        cov = "\n".join(str(cell.value) for row in wb["Coverage"].iter_rows()
                        for cell in row if cell.value is not None)
        for spec in modmod.all_modules():
            assert spec.title in cov, f"{spec.slug} missing from coverage"
    finally:
        modmod.MODULES.clear()
        modmod.MODULES.update(originals)
