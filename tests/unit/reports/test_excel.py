"""Renderer tests: build_report over a ReportModel and the legacy dict."""
import io

import openpyxl

from ruckus_dashboard.reports.excel import build_report, _safe_sheet_name
from ruckus_dashboard.reports.model import (
    ColumnSpec, DrillSample, ModuleReport, ReportModel,
)


def _model():
    return ReportModel(
        generated_at="2026-06-30T07:00:00Z", connection_label="SZ-LAB",
        modules=[
            ModuleReport(
                slug="aps", title="Access Points", group="Wireless", status="ok",
                columns=[ColumnSpec("Name", "name"),
                         ColumnSpec("Status", "status", "status")],
                summary={"total": 2, "online": 1, "offline": 1},
                rows=[{"id": "a", "name": "AP1", "status": "online"},
                      {"id": "b", "name": "AP2", "status": "offline"}],
                row_total=2,
                raw_samples=[{"apMac": "a", "deviceName": "AP1"}],
                drill_samples=[DrillSample("a", {"identity": {"name": "AP1"}})],
            ),
            ModuleReport(slug="topology", title="Topology", group="Cross-cutting",
                         status="disabled", note="module unavailable"),
        ],
    )


def test_safe_sheet_name_truncates_and_dedupes():
    used: set[str] = set()
    a = _safe_sheet_name("A very long module title that exceeds excel limit", used)
    assert len(a) <= 31 and a not in ("",)
    used.add(a)
    b = _safe_sheet_name("A very long module title that exceeds excel limit", used)
    assert b != a and len(b) <= 31           # deduped suffix


def test_safe_sheet_name_strips_illegal_chars():
    used: set[str] = set()
    name = _safe_sheet_name("APs: by/zone [x]?", used)
    for bad in ":/\\?*[]":
        assert bad not in name


def test_build_report_from_model_loads_with_overview_and_coverage():
    blob = build_report(_model())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert "Overview" in wb.sheetnames
    assert "Coverage" in wb.sheetnames
    # Coverage lists every module + its status.
    cov_text = "\n".join(str(c.value) for row in wb["Coverage"].iter_rows()
                         for c in row if c.value is not None)
    assert "Access Points" in cov_text and "Topology" in cov_text
    assert "disabled" in cov_text


def test_build_report_legacy_dict_still_renders_curated_sheets():
    data = {
        "aps": [{"name": "AP1", "zone": "HQ", "status": "online", "mac": "a"},
                {"name": "AP2", "zone": "HQ", "status": "offline", "mac": "b"}],
        "clients": [{"hostname": "h1", "mac": "m", "ssid": "S", "ap": "AP1",
                     "band": "5 GHz", "quality": "good",
                     "rx_bytes": 10, "tx_bytes": 20}],
        "alarms": [{"severity": "critical", "category": "AP", "source": "AP2",
                    "message": "down", "count": 1}],
        "switches": [{"name": "SW1", "ip": "10.0.0.1", "model": "ICX",
                      "fw": "x", "status": "online", "ports_online": 10,
                      "ports_total": 24, "group": "Core", "mac": "c"}],
    }
    wb = openpyxl.load_workbook(io.BytesIO(build_report(data)))
    # Curated sheets preserved; charts intact (regression for current suite).
    assert {"Overview", "APs by Zone", "Clients", "Alarms",
            "Switches", "Offline Devices"} <= set(wb.sheetnames)
    assert len(wb["APs by Zone"]._charts) == 1
    assert len(wb["Clients"]._charts) == 1
    assert len(wb["Alarms"]._charts) == 1


def test_overview_alarm_count_uses_or_zero_for_missing_or_zero_count():
    # Parity with scheduler.py #14: an alarm whose "count" is missing/0 must
    # contribute 0 (not a phantom 1) to the Active/Critical alarm totals.
    data = {
        "aps": [], "clients": [], "switches": [],
        "alarms": [
            {"severity": "critical"},              # missing count -> 0
            {"severity": "critical", "count": 0},  # explicit 0    -> 0
            {"severity": "critical", "count": 3},  # counts        -> 3
        ],
    }
    ws = openpyxl.load_workbook(io.BytesIO(build_report(data)))["Overview"]
    overview = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                for r in range(5, ws.max_row + 1)}
    assert overview["Active alarms"] == 3
    assert overview["Critical alarms"] == 3


def test_module_sheet_has_summary_list_raw_and_drill():
    blob = build_report(_model())
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    # The aps module gets its own sheet (safe name == title, fits in 31 chars).
    assert "Access Points" in wb.sheetnames
    ws = wb["Access Points"]
    text = "\n".join(str(c.value) for row in ws.iter_rows()
                     for c in row if c.value is not None)
    assert "Summary" in text and "online" in text      # summary block
    assert "AP1" in text and "AP2" in text              # list rows
    assert "apMac" in text                              # raw field-map key
    assert "Drill" in text                              # drill block label


def test_disabled_module_sheet_notes_status():
    wb = openpyxl.load_workbook(io.BytesIO(build_report(_model())))
    assert "Topology" in wb.sheetnames
    ws = wb["Topology"]
    text = "\n".join(str(c.value) for row in ws.iter_rows()
                     for c in row if c.value is not None)
    assert "disabled" in text


def test_list_rows_capped_with_more_note():
    big = ReportModel(
        generated_at="t", connection_label="x",
        modules=[ModuleReport(
            slug="clients", title="Clients", group="Wireless", status="ok",
            columns=[ColumnSpec("Host", "hostname")],
            rows=[{"id": str(i), "hostname": f"h{i}"} for i in range(1500)],
            row_total=1500)])
    wb = openpyxl.load_workbook(io.BytesIO(build_report(big)))
    # The generic clients module sheet (title "Clients") is deduped against the
    # curated "Clients" chart sheet — locate it by its A1 heading, not by name.
    ws = next(s for s in wb.worksheets if s["A1"].value == "Clients"
              and s["A2"].value == "Status: ok")
    text = "\n".join(str(c.value) for row in ws.iter_rows()
                     for c in row if c.value is not None)
    assert "more" in text.lower()                       # "+N more" note present
