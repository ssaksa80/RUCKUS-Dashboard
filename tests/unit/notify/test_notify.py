"""Unit tests for notification config, rules, mailer, scheduler due-logic,
and the Excel report builder."""
import json
import os
import stat
import sys
import time

import pytest

from ruckus_dashboard.notify import config as cfg_mod
from ruckus_dashboard.notify.config import save_config, _path
from ruckus_dashboard.notify.rules import evaluate
from ruckus_dashboard.notify.scheduler import NotifyScheduler, state_from_data
from ruckus_dashboard.reports.excel import build_report


class FakeSecrets:
    def encrypt(self, s):
        return f"enc:{s}"

    def decrypt(self, s):
        assert s.startswith("enc:")
        return s[4:]


# ── config ───────────────────────────────────────────────────────────────

def test_notifications_file_is_chmod_600(tmp_path):
    class _Sec:
        def encrypt(self, s): return "enc:" + s
    save_config(str(tmp_path), {"smtp": {"password": "pw"}}, _Sec())
    p = _path(str(tmp_path))
    assert p.exists()
    if sys.platform != "win32":
        assert stat.S_IMODE(os.stat(p).st_mode) == 0o600


def test_config_password_encrypted_masked_and_preserved(tmp_path):
    secrets = FakeSecrets()
    saved = cfg_mod.save_config(str(tmp_path), {
        "smtp": {"host": "mail.x", "password": "hunter2"},
        "alerts": {"enabled": True, "recipients": ["a@x"]},
    }, secrets)
    # encrypted at rest, never plaintext
    raw = json.loads((tmp_path / "notifications.json").read_text())
    assert raw["smtp"]["password_enc"] == "enc:hunter2"
    assert "hunter2" not in json.dumps(raw["smtp"].get("password", ""))
    # masked for display
    disp = cfg_mod.display_config(saved)
    assert disp["smtp"]["password"] == cfg_mod.PASSWORD_MASK
    # posting the mask back preserves the stored secret
    saved2 = cfg_mod.save_config(str(tmp_path), {
        "smtp": {"host": "mail.x", "password": cfg_mod.PASSWORD_MASK},
    }, secrets)
    assert cfg_mod.smtp_password(saved2, secrets) == "hunter2"


def test_config_defaults_when_missing(tmp_path):
    cfg = cfg_mod.load_config(str(tmp_path))
    assert cfg["smtp"]["port"] == 587
    assert cfg["alerts"]["rules"]["critical_alarm"] is True
    assert cfg["report"]["time"] == "07:00"


class _Sec2:
    def encrypt(self, s): return "enc:" + s
    def decrypt(self, s): return s


def test_partial_post_preserves_other_subkeys(tmp_path):
    save_config(str(tmp_path), {"report": {"enabled": True, "recipients": ["a@x"], "time": "06:00"}}, _Sec2())
    save_config(str(tmp_path), {"report": {"enabled": False}}, _Sec2())
    cfg = cfg_mod.load_config(str(tmp_path))
    assert cfg["report"]["enabled"] is False
    assert cfg["report"]["recipients"] == ["a@x"]
    assert cfg["report"]["time"] == "06:00"


# ── rules ────────────────────────────────────────────────────────────────

def test_rules_fire_on_transition_only():
    rules = {"ap_offline": True, "switch_offline": True, "critical_alarm": True}
    first = evaluate(None, {"aps_offline": 2}, rules, 1)
    assert len(first) == 1 and "2" in first[0]
    # unchanged level → silence
    again = evaluate({"aps_offline": 2}, {"aps_offline": 2}, rules, 1)
    assert again == []
    # rises again → fires
    more = evaluate({"aps_offline": 2}, {"aps_offline": 5}, rules, 1)
    assert len(more) == 1


def test_rules_threshold_and_toggles():
    rules = {"ap_offline": True, "switch_offline": False, "critical_alarm": True}
    assert evaluate({}, {"aps_offline": 1}, rules, 3) == []     # below threshold
    assert evaluate({}, {"switches_offline": 9}, rules, 1) == []  # rule off
    crit = evaluate({}, {"critical_alarms": 1}, rules, 1)
    assert len(crit) == 1


# ── scheduler due-logic ──────────────────────────────────────────────────

def _sched(tmp_path):
    return NotifyScheduler(str(tmp_path), {}, FakeSecrets())


def test_alerts_due_respects_interval(tmp_path):
    s = _sched(tmp_path)
    cfg = cfg_mod.load_config(str(tmp_path))
    cfg["alerts"]["enabled"] = True
    cfg["alerts"]["check_seconds"] = 300
    assert s._alerts_due(cfg, time.time()) is True
    s._last_alert_check = time.time()
    assert s._alerts_due(cfg, time.time()) is False
    cfg["alerts"]["enabled"] = False
    assert s._alerts_due(cfg, time.time() + 9999) is False


def test_report_due_once_per_day_after_time(tmp_path):
    s = _sched(tmp_path)
    cfg = cfg_mod.load_config(str(tmp_path))
    cfg["report"]["enabled"] = True
    cfg["report"]["time"] = "07:00"
    before = time.strptime("2026-06-10 06:59", "%Y-%m-%d %H:%M")
    after = time.strptime("2026-06-10 07:01", "%Y-%m-%d %H:%M")
    assert s._report_due(cfg, before) is False
    assert s._report_due(cfg, after) is True
    s._last_report_day = "2026-06-10"
    assert s._report_due(cfg, after) is False      # once per day
    next_day = time.strptime("2026-06-11 07:01", "%Y-%m-%d %H:%M")
    assert s._report_due(cfg, next_day) is True


def test_state_from_data_counts():
    data = {"aps": [{"status": "offline"}, {"status": "online"}],
            "switches": [{"status": "offline"}],
            "alarms": [{"severity": "critical", "count": 3},
                       {"severity": "major", "count": 1}]}
    s = state_from_data(data)
    assert s["aps_offline"] == 1
    assert s["switches_offline"] == 1
    assert s["critical_alarms"] == 3
    assert s["poor_aps"] == []


def test_poor_quality_aps_threshold():
    from ruckus_dashboard.notify.scheduler import poor_quality_aps
    clients = ([{"ap": "AP-BAD", "quality": "poor"}] * 4 +
               [{"ap": "AP-BAD", "quality": "good"}] +     # 4/5 poor = 80%
               [{"ap": "AP-OK", "quality": "poor"}] +      # only 1 client -> skip
               [{"ap": "AP-FINE", "quality": "good"}] * 5)
    flagged = poor_quality_aps(clients)
    assert flagged == ["AP-BAD (4/5 poor)"]


def test_poor_ap_rule_fires_for_new_aps_only():
    rules = {"poor_client_ap": True}
    first = evaluate(None, {"poor_aps": ["AP-1 (4/5 poor)"]}, rules, 1)
    assert len(first) == 1 and "AP-1" in first[0]
    same = evaluate({"poor_aps": ["AP-1 (4/5 poor)"]},
                    {"poor_aps": ["AP-1 (5/6 poor)"]}, rules, 1)
    assert same == []                       # same AP still degraded -> silent
    new = evaluate({"poor_aps": ["AP-1 (4/5 poor)"]},
                   {"poor_aps": ["AP-1 (4/5 poor)", "AP-2 (3/3 poor)"]}, rules, 1)
    assert len(new) == 1 and "AP-2" in new[0]


# ── excel report ─────────────────────────────────────────────────────────

def test_build_report_loads_and_has_sheets_and_charts():
    import openpyxl
    import io
    data = {
        "aps": [{"name": "AP1", "zone": "HQ", "status": "online", "mac": "a"},
                {"name": "AP2", "zone": "HQ", "status": "offline", "mac": "b"}],
        "clients": [{"hostname": "h1", "mac": "m", "ssid": "S", "ap": "AP1",
                     "band": "5 GHz", "quality": "good",
                     "rx_bytes": 10, "tx_bytes": 20}],
        "alarms": [{"severity": "critical", "category": "AP", "source": "AP2",
                    "message": "down", "count": 1}],
        "switches": [{"name": "SW1", "ip": "10.0.0.1", "model": "ICX",
                      "fw": "x", "status": "online", "ports_online": 10,
                      "ports_total": 24, "group": "Core", "mac": "c"}],
    }
    blob = build_report(data)
    wb = openpyxl.load_workbook(io.BytesIO(blob))
    assert set(wb.sheetnames) == {"Overview", "APs by Zone", "Clients",
                                  "Alarms", "Switches", "Offline Devices"}
    assert len(wb["APs by Zone"]._charts) == 1   # bar
    assert len(wb["Clients"]._charts) == 1       # pie
    assert len(wb["Alarms"]._charts) == 1        # pie
    # offline AP listed
    assert any(c.value == "AP2" for row in wb["Offline Devices"].iter_rows()
               for c in row)


# ── mailer (monkeypatched smtplib) ───────────────────────────────────────

def test_send_email_via_monkeypatched_smtp(monkeypatch):
    from ruckus_dashboard.notify import mailer
    sent = {}

    class FakeSMTP:
        def __init__(self, host, port, timeout=0):
            sent["host"], sent["port"] = host, port

        def __enter__(self): return self
        def __exit__(self, *a): return False
        def ehlo(self): sent["ehlo"] = sent.get("ehlo", 0) + 1
        def starttls(self): sent["tls"] = True
        def login(self, u, p): sent["login"] = (u, p)
        def send_message(self, msg): sent["subject"] = msg["Subject"]

    monkeypatch.setattr(mailer.smtplib, "SMTP", FakeSMTP)
    cfg = {"smtp": {"host": "mail.x", "port": 587, "use_tls": True,
                    "username": "u", "from_addr": "dso@x"}}
    mailer.send_email(cfg, "pw", ["a@x"], "Subject!", "body")
    assert sent["host"] == "mail.x" and sent["tls"] is True
    assert sent["ehlo"] == 2   # before and after STARTTLS (networker pattern)
    assert sent["login"] == ("u", "pw")
    assert sent["subject"] == "Subject!"


def test_send_email_requires_host_and_recipients():
    from ruckus_dashboard.notify.mailer import send_email
    with pytest.raises(ValueError):
        send_email({"smtp": {"host": ""}}, "", ["a@x"], "s", "b")
    with pytest.raises(ValueError):
        send_email({"smtp": {"host": "mail.x"}}, "", [], "s", "b")


def test_send_email_ssl_mode_uses_smtp_ssl(monkeypatch):
    from ruckus_dashboard.notify import mailer
    used = {}

    class FakeSSL:
        def __init__(self, host, port, timeout=0): used["ssl"] = (host, port)
        def __enter__(self): return self
        def __exit__(self, *a): return False
        def login(self, u, p): used["login"] = (u, p)
        def send_message(self, msg): used["sent"] = True

    monkeypatch.setattr(mailer.smtplib, "SMTP_SSL", FakeSSL)
    cfg = {"smtp": {"host": "mail.x", "port": 465, "security": "ssl",
                    "username": "u", "from_addr": "dso@x"}}
    out = mailer.send_email(cfg, "pw", ["a@x"], "s", "b")
    assert used["ssl"] == ("mail.x", 465) and used["sent"]
    assert out["stage"] == "sent" and out["security"] == "ssl"


def test_send_email_reports_failing_stage(monkeypatch):
    import smtplib as real_smtplib
    from ruckus_dashboard.notify import mailer

    class FakeSMTP:
        def __init__(self, host, port, timeout=0): pass
        def __enter__(self): return self
        def __exit__(self, *a): return False
        def ehlo(self): pass
        def starttls(self): pass
        def login(self, u, p):
            raise real_smtplib.SMTPAuthenticationError(535, b"5.7.8 Bad creds")
        def send_message(self, msg): pass

    monkeypatch.setattr(mailer.smtplib, "SMTP", FakeSMTP)
    cfg = {"smtp": {"host": "mail.x", "port": 587, "security": "starttls",
                    "username": "u"}}
    with pytest.raises(mailer.SmtpDeliveryError) as ei:
        mailer.send_email(cfg, "wrong", ["a@x"], "s", "b")
    assert ei.value.stage == "login"
    assert "authentication rejected" in ei.value.detail
    assert "5.7.8" in ei.value.detail


def test_traffic_live_rate_from_deltas(monkeypatch):
    """Second traffic poll derives bits/s from the cumulative-byte delta."""
    import responses as resp
    import time as _t
    from ruckus_dashboard.modules import traffic as traffic_mod
    from ruckus_dashboard.auth.session_store import ConnectionConfig
    from ruckus_dashboard.modules._base import FetcherContext
    from ruckus_dashboard.infra.capability_gate import CapabilityGate

    cfg = {"RUCKUS_TIMEOUT_SECONDS": 5, "RUCKUS_DEBUG_BYTES": 1000,
           "RUCKUS_PAGE_LIMIT": 500, "RUCKUS_HOST_ALLOWLIST": None}
    conn = ConnectionConfig(platform="smartzone",
        api_base="https://sz.example:8443/wsg/api/public", display_name="SZ",
        auth_token="t", api_version="v11_0", verify_tls=False,
        token_expires_at=9999999999)
    ctx = FetcherContext(connection=conn, config=cfg, filters=None,
        capability_gate=CapabilityGate(set()), connection_label="SZ")

    sw = "https://sz.example:8443/switchm/api"
    traffic_mod._PREV["base"] = {}
    traffic_mod._PREV["rate"] = {}

    with resp.RequestsMock() as rs:
        rs.add(resp.POST, f"{sw}/v11_0/switch",
               json={"list": [{"id": "S1", "switchName": "SW-1"}],
                     "totalCount": 1, "hasMore": False}, status=200)
        rs.add(resp.POST, f"{sw}/v11_0/traffic/top/usage",
               json={"list": [{"key": "S1", "value": 1000}]}, status=200)
        out1 = traffic_mod.fetch(ctx)
    assert out1["items"][0]["rate_bps"] is None    # first sample

    # Pretend the baseline was taken 10 s ago.
    traffic_mod._PREV["base"]["S1"]["t"] = _t.time() - 10

    with resp.RequestsMock() as rs:
        rs.add(resp.POST, f"{sw}/v11_0/switch",
               json={"list": [{"id": "S1", "switchName": "SW-1"}],
                     "totalCount": 1, "hasMore": False}, status=200)
        rs.add(resp.POST, f"{sw}/v11_0/traffic/top/usage",
               json={"list": [{"key": "S1", "value": 11000}]}, status=200)
        out2 = traffic_mod.fetch(ctx)
    rate = out2["items"][0]["rate_bps"]
    assert rate is not None and 7000 <= rate <= 9000   # ~8000 bps

    # Third poll with an UNCHANGED counter keeps the last rate (the SmartZone
    # aggregate refreshes periodically; a flat counter must not read 0 bps).
    with resp.RequestsMock() as rs:
        rs.add(resp.POST, f"{sw}/v11_0/switch",
               json={"list": [{"id": "S1", "switchName": "SW-1"}],
                     "totalCount": 1, "hasMore": False}, status=200)
        rs.add(resp.POST, f"{sw}/v11_0/traffic/top/usage",
               json={"list": [{"key": "S1", "value": 11000}]}, status=200)
        out3 = traffic_mod.fetch(ctx)
    assert out3["items"][0]["rate_bps"] == rate


# ── outage: DeviceStatus and device_online helpers ────────────────────────

def test_device_status_dataclass_fields():
    from ruckus_dashboard.notify.outage import DeviceStatus
    ds = DeviceStatus(
        key="ap:aabbcc",
        type="ap",
        name="AP-1",
        group="HQ",
        online=True,
        raw_status="online",
        last_change=1000.0,
    )
    assert ds.key == "ap:aabbcc"
    assert ds.online is True
    assert ds.last_change == 1000.0
    # pending fields default to None
    assert ds.pending_since is None
    assert ds.pending_target is None


def test_device_online_ap():
    from ruckus_dashboard.notify.outage import device_online
    assert device_online("ap", "online") is True
    assert device_online("ap", "offline") is False
    assert device_online("ap", "unknown") is False


def test_device_online_switch():
    from ruckus_dashboard.notify.outage import device_online
    assert device_online("switch", "online") is True
    assert device_online("switch", "offline") is False
    assert device_online("switch", "flagged") is False


def test_device_online_controller():
    from ruckus_dashboard.notify.outage import device_online
    # matches controller._NODE_ONLINE exactly
    for state in ("in_service", "online", "active", "up",
                  "management_in_service", "service_ready"):
        assert device_online("controller", state) is True, state
    assert device_online("controller", "disconnected") is False
    assert device_online("controller", "") is False


# ── outage: OutageEngine.reconcile ────────────────────────────────────────

def _make_snapshot(entries: list[tuple]) -> dict:
    """entries: (key, type, name, group, online, raw_status)"""
    from ruckus_dashboard.notify.outage import DeviceStatus
    return {
        key: DeviceStatus(key=key, type=typ, name=name, group=grp,
                          online=on, raw_status=rs, last_change=0.0)
        for key, typ, name, grp, on, rs in entries
    }


def test_reconcile_baseline_seeding_no_events():
    """First call (empty prior devices) seeds silently — no events emitted."""
    from ruckus_dashboard.notify.outage import OutageEngine
    snapshot = _make_snapshot([
        ("ap:aa", "ap", "AP-1", "HQ", False, "offline"),
        ("ap:bb", "ap", "AP-2", "HQ", True, "online"),
    ])
    cfg = {"debounce_seconds": 0, "recovery": True, "offline_threshold": 1}
    events, new_devices = OutageEngine.reconcile({}, snapshot, cfg, now=1000.0)
    assert events == []
    assert new_devices["ap:aa"].online is False
    assert new_devices["ap:bb"].online is True


def test_reconcile_offline_transition_fires_immediately_when_no_debounce():
    """Previously-online device goes offline; debounce=0 fires on first tick."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev_devices = _make_snapshot([
        ("ap:aa", "ap", "AP-1", "HQ", True, "online"),
    ])
    snapshot = _make_snapshot([
        ("ap:aa", "ap", "AP-1", "HQ", False, "offline"),
    ])
    cfg = {"debounce_seconds": 0, "recovery": True, "offline_threshold": 1}
    events, _ = OutageEngine.reconcile(prev_devices, snapshot, cfg, now=2000.0)
    assert len(events) == 1
    assert events[0].kind == "offline"
    assert events[0].key == "ap:aa"
    assert events[0].name == "AP-1"
    assert events[0].group == "HQ"
    assert events[0].ts == 2000.0


def test_reconcile_stable_online_no_event():
    """Device that was online and stays online emits nothing."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", True, "online")])
    snap = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", True, "online")])
    cfg = {"debounce_seconds": 0, "recovery": True, "offline_threshold": 1}
    events, _ = OutageEngine.reconcile(prev, snap, cfg, now=3000.0)
    assert events == []


def test_reconcile_stable_offline_no_event():
    """Pre-existing outage (committed offline) stays offline — no re-alert."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", False, "offline")])
    snap = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", False, "offline")])
    cfg = {"debounce_seconds": 0, "recovery": True, "offline_threshold": 1}
    events, _ = OutageEngine.reconcile(prev, snap, cfg, now=4000.0)
    assert events == []


def test_reconcile_recovery_event_when_enabled():
    """Offline device comes back; recovery=True emits online event."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev = _make_snapshot([("sw:s1", "switch", "SW-1", "Core", False, "offline")])
    snap = _make_snapshot([("sw:s1", "switch", "SW-1", "Core", True, "online")])
    cfg = {"debounce_seconds": 0, "recovery": True, "offline_threshold": 1}
    events, _ = OutageEngine.reconcile(prev, snap, cfg, now=5000.0)
    assert len(events) == 1
    assert events[0].kind == "online"
    assert events[0].key == "sw:s1"


def test_reconcile_recovery_suppressed_when_disabled():
    """Offline device comes back; recovery=False emits nothing."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev = _make_snapshot([("sw:s1", "switch", "SW-1", "Core", False, "offline")])
    snap = _make_snapshot([("sw:s1", "switch", "SW-1", "Core", True, "online")])
    cfg = {"debounce_seconds": 0, "recovery": False, "offline_threshold": 1}
    events, _ = OutageEngine.reconcile(prev, snap, cfg, now=6000.0)
    assert events == []


def test_reconcile_offline_threshold_suppresses_small_batch():
    """`offline_threshold=3` suppresses a batch of only 2 newly-offline devices."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev = _make_snapshot([
        ("ap:a1", "ap", "AP-1", "HQ", True, "online"),
        ("ap:a2", "ap", "AP-2", "HQ", True, "online"),
    ])
    snap = _make_snapshot([
        ("ap:a1", "ap", "AP-1", "HQ", False, "offline"),
        ("ap:a2", "ap", "AP-2", "HQ", False, "offline"),
    ])
    cfg = {"debounce_seconds": 0, "recovery": True, "offline_threshold": 3}
    events, _ = OutageEngine.reconcile(prev, snap, cfg, now=7000.0)
    # only 2 newly offline, threshold=3 — suppressed
    offline_events = [e for e in events if e.kind == "offline"]
    assert offline_events == []


def test_reconcile_offline_threshold_fires_when_met():
    """`offline_threshold=2` fires when exactly 2 devices newly go offline."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev = _make_snapshot([
        ("ap:a1", "ap", "AP-1", "HQ", True, "online"),
        ("ap:a2", "ap", "AP-2", "HQ", True, "online"),
    ])
    snap = _make_snapshot([
        ("ap:a1", "ap", "AP-1", "HQ", False, "offline"),
        ("ap:a2", "ap", "AP-2", "HQ", False, "offline"),
    ])
    cfg = {"debounce_seconds": 0, "recovery": True, "offline_threshold": 2}
    events, _ = OutageEngine.reconcile(prev, snap, cfg, now=8000.0)
    offline_events = [e for e in events if e.kind == "offline"]
    assert len(offline_events) == 2


# ── outage: debounce ──────────────────────────────────────────────────────

def test_reconcile_debounce_holds_on_first_tick():
    """Offline device within debounce window: no event on first tick."""
    from ruckus_dashboard.notify.outage import OutageEngine
    prev = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", True, "online")])
    snap = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", False, "offline")])
    cfg = {"debounce_seconds": 120, "recovery": True, "offline_threshold": 1}
    events, new_devices = OutageEngine.reconcile(prev, snap, cfg, now=1000.0)
    assert events == []
    # pending_since recorded
    assert new_devices["ap:aa"].pending_since == 1000.0
    assert new_devices["ap:aa"].pending_target is False
    # committed state unchanged
    assert new_devices["ap:aa"].online is True


def test_reconcile_debounce_fires_after_window():
    """Same device still offline after debounce_seconds → event emitted."""
    from ruckus_dashboard.notify.outage import OutageEngine, DeviceStatus
    # Simulate prev_devices as the state AFTER the first tick (pending set).
    prev_ds = DeviceStatus(
        key="ap:aa", type="ap", name="AP-1", group="HQ",
        online=True, raw_status="online", last_change=900.0,
        pending_since=1000.0, pending_target=False,
    )
    snap = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", False, "offline")])
    cfg = {"debounce_seconds": 120, "recovery": True, "offline_threshold": 1}
    events, new_devices = OutageEngine.reconcile(
        {"ap:aa": prev_ds}, snap, cfg, now=1121.0  # 121 s later, past window
    )
    assert len(events) == 1
    assert events[0].kind == "offline"
    assert new_devices["ap:aa"].online is False
    assert new_devices["ap:aa"].pending_since is None


def test_reconcile_debounce_not_yet_matured():
    """Still within the window (119 s): pending kept, no event."""
    from ruckus_dashboard.notify.outage import OutageEngine, DeviceStatus
    prev_ds = DeviceStatus(
        key="ap:aa", type="ap", name="AP-1", group="HQ",
        online=True, raw_status="online", last_change=900.0,
        pending_since=1000.0, pending_target=False,
    )
    snap = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", False, "offline")])
    cfg = {"debounce_seconds": 120, "recovery": True, "offline_threshold": 1}
    events, new_devices = OutageEngine.reconcile(
        {"ap:aa": prev_ds}, snap, cfg, now=1119.0
    )
    assert events == []
    assert new_devices["ap:aa"].pending_since == 1000.0
    assert new_devices["ap:aa"].online is True


def test_reconcile_flap_within_debounce_suppressed():
    """Device goes offline then recovers within the debounce window — no event."""
    from ruckus_dashboard.notify.outage import OutageEngine, DeviceStatus
    # After tick 1: pending toward False (offline)
    prev_ds = DeviceStatus(
        key="ap:aa", type="ap", name="AP-1", group="HQ",
        online=True, raw_status="online", last_change=900.0,
        pending_since=1000.0, pending_target=False,
    )
    # Tick 2: device is back online within the 120 s window.
    snap = _make_snapshot([("ap:aa", "ap", "AP-1", "HQ", True, "online")])
    cfg = {"debounce_seconds": 120, "recovery": True, "offline_threshold": 1}
    events, new_devices = OutageEngine.reconcile(
        {"ap:aa": prev_ds}, snap, cfg, now=1060.0
    )
    assert events == []
    # Pending cleared; stable online.
    assert new_devices["ap:aa"].pending_since is None
    assert new_devices["ap:aa"].online is True


# ── outage: render_alert ──────────────────────────────────────────────────

def _make_event(kind, key, typ, name, group, ts=5000.0):
    from ruckus_dashboard.notify.outage import OutageEvent
    return OutageEvent(kind=kind, key=key, type=typ, name=name,
                       group=group, raw_status="offline", ts=ts)


def test_render_alert_subject_counts():
    from ruckus_dashboard.notify.outage import render_alert
    events = [
        _make_event("offline", "ap:a1", "ap", "AP-1", "HQ"),
        _make_event("offline", "ap:a2", "ap", "AP-2", "Branch"),
        _make_event("online",  "sw:s1", "switch", "SW-1", "Core"),
    ]
    note = render_alert(events, group_by="site")
    assert "2 devices offline" in note.subject
    assert "1 recovered" in note.subject


def test_render_alert_subject_includes_groups():
    from ruckus_dashboard.notify.outage import render_alert
    events = [
        _make_event("offline", "ap:a1", "ap", "AP-1", "HQ"),
        _make_event("offline", "ap:a2", "ap", "AP-2", "Branch"),
    ]
    note = render_alert(events, group_by="site")
    assert "HQ" in note.subject or "Branch" in note.subject


def test_render_alert_body_groups_by_site():
    from ruckus_dashboard.notify.outage import render_alert
    events = [
        _make_event("offline", "ap:a1", "ap", "AP-1", "HQ"),
        _make_event("offline", "sw:s1", "switch", "SW-1", "HQ"),
        _make_event("offline", "ap:a2", "ap", "AP-2", "Branch"),
    ]
    note = render_alert(events, group_by="site")
    assert "HQ" in note.body
    assert "Branch" in note.body
    assert "AP-1" in note.body
    assert "SW-1" in note.body
    assert "AP-2" in note.body


def test_render_alert_body_flat_when_group_by_none():
    from ruckus_dashboard.notify.outage import render_alert
    events = [
        _make_event("offline", "ap:a1", "ap", "AP-1", "HQ"),
        _make_event("offline", "ap:a2", "ap", "AP-2", "Branch"),
    ]
    note = render_alert(events, group_by="none")
    assert "AP-1" in note.body
    assert "AP-2" in note.body


def test_render_alert_recovery_section_in_body():
    from ruckus_dashboard.notify.outage import render_alert
    events = [
        _make_event("offline", "ap:a1", "ap", "AP-1", "HQ"),
        _make_event("online",  "sw:s1", "switch", "SW-1", "Core"),
    ]
    note = render_alert(events, group_by="site")
    assert "Recovered" in note.body or "recovered" in note.body
    assert "SW-1" in note.body


def test_render_alert_structured_events_tuple():
    from ruckus_dashboard.notify.outage import render_alert
    events = [_make_event("offline", "ap:a1", "ap", "AP-1", "HQ")]
    note = render_alert(events, group_by="site")
    assert len(note.events) == 1
    assert note.events[0].key == "ap:a1"


def test_render_alert_controller_node_event():
    from ruckus_dashboard.notify.outage import render_alert
    events = [_make_event("offline", "controller:node1",
                          "controller", "SZ-Node-1", "controller")]
    note = render_alert(events, group_by="site")
    assert "SZ-Node-1" in note.body
    assert "controller" in note.body.lower()


# ── state_store ───────────────────────────────────────────────────────────

def test_json_state_store_roundtrip(tmp_instance):
    from ruckus_dashboard.notify.state_store import JsonOutageStateStore
    from ruckus_dashboard.notify.outage import DeviceStatus
    store = JsonOutageStateStore(tmp_instance)
    ds = DeviceStatus(key="ap:aa", type="ap", name="AP-1", group="HQ",
                      online=False, raw_status="offline", last_change=1000.0)
    state = {
        "devices": {"ap:aa": ds},
        "report": {"last_report_day": "2026-06-30"},
    }
    store.save(state)
    loaded = store.load()
    assert loaded["devices"]["ap:aa"].online is False
    assert loaded["devices"]["ap:aa"].name == "AP-1"
    assert loaded["report"]["last_report_day"] == "2026-06-30"


def test_json_state_store_missing_file_returns_empty(tmp_instance):
    from ruckus_dashboard.notify.state_store import JsonOutageStateStore
    store = JsonOutageStateStore(tmp_instance)
    result = store.load()
    assert result["devices"] == {}
    assert result["report"] == {}


def test_json_state_store_corrupt_file_returns_empty(tmp_instance):
    from pathlib import Path
    from ruckus_dashboard.notify.state_store import JsonOutageStateStore
    p = Path(tmp_instance) / "notify_state.json"
    p.write_text("NOT JSON{{{", encoding="utf-8")
    store = JsonOutageStateStore(tmp_instance)
    result = store.load()
    assert result["devices"] == {}


def test_json_state_store_atomic_no_tmp_left(tmp_instance):
    """After save(), .tmp file must not exist."""
    from pathlib import Path
    from ruckus_dashboard.notify.state_store import JsonOutageStateStore
    store = JsonOutageStateStore(tmp_instance)
    store.save({"devices": {}, "report": {}})
    tmp = Path(tmp_instance) / "notify_state.json.tmp"
    assert not tmp.exists()
    main = Path(tmp_instance) / "notify_state.json"
    assert main.exists()


def test_json_state_store_chmod_best_effort_on_windows(tmp_instance):
    """chmod failure (Windows) must not raise."""
    from unittest.mock import patch
    from ruckus_dashboard.notify.state_store import JsonOutageStateStore
    store = JsonOutageStateStore(tmp_instance)
    with patch("os.chmod", side_effect=OSError("read-only fs")):
        # Should not raise despite chmod failing.
        store.save({"devices": {}, "report": {}})


def test_json_state_store_pending_fields_roundtrip(tmp_instance):
    """pending_since and pending_target survive a save/load cycle."""
    from ruckus_dashboard.notify.state_store import JsonOutageStateStore
    from ruckus_dashboard.notify.outage import DeviceStatus
    store = JsonOutageStateStore(tmp_instance)
    ds = DeviceStatus(key="ap:bb", type="ap", name="AP-2", group="Branch",
                      online=True, raw_status="online", last_change=900.0,
                      pending_since=1000.0, pending_target=False)
    store.save({"devices": {"ap:bb": ds}, "report": {}})
    loaded = store.load()
    assert loaded["devices"]["ap:bb"].pending_since == 1000.0
    assert loaded["devices"]["ap:bb"].pending_target is False


# ── channels ─────────────────────────────────────────────────────────────

def test_notification_is_importable_from_channels():
    from ruckus_dashboard.notify.channels import Notification
    from ruckus_dashboard.notify.outage import Notification as NotifOutage
    # channels re-exports the same class from outage.py
    assert Notification is NotifOutage


def test_email_channel_is_configured_when_recipients():
    from ruckus_dashboard.notify.channels import EmailChannel
    ch = EmailChannel()
    assert ch.name == "email"
    assert ch.is_configured({"alerts": {"recipients": ["a@x"]}}) is True
    assert ch.is_configured({"alerts": {"recipients": []}}) is False
    assert ch.is_configured({}) is False


def test_email_channel_send_calls_mailer(monkeypatch):
    from ruckus_dashboard.notify import channels as ch_mod
    from ruckus_dashboard.notify.channels import EmailChannel
    from ruckus_dashboard.notify.outage import Notification
    calls = {}

    def fake_send(cfg, pw, recipients, subject, body, **kw):
        calls["recipients"] = recipients
        calls["subject"] = subject
        calls["body"] = body

    monkeypatch.setattr(ch_mod, "send_email", fake_send)
    monkeypatch.setattr(ch_mod, "smtp_password", lambda cfg, secrets: "pw")

    note = Notification(
        subject="[RUCKUS DSO] 1 device offline (HQ)",
        body="DEVICES OFFLINE\n  AP-1 (ap) — offline",
        events=(),
    )
    cfg = {"alerts": {"recipients": ["noc@x"]}, "smtp": {"host": "mail.x"}}
    EmailChannel().send(cfg, object(), note)
    assert calls["recipients"] == ["noc@x"]
    assert "1 device offline" in calls["subject"]


def test_email_channel_send_failure_does_not_raise(monkeypatch):
    """A raising mailer.send_email is swallowed by the channel."""
    from ruckus_dashboard.notify import channels as ch_mod
    from ruckus_dashboard.notify.channels import EmailChannel
    from ruckus_dashboard.notify.outage import Notification

    monkeypatch.setattr(ch_mod, "send_email",
                        lambda *a, **kw: (_ for _ in ()).throw(RuntimeError("SMTP down")))
    monkeypatch.setattr(ch_mod, "smtp_password", lambda cfg, secrets: "")
    note = Notification(subject="s", body="b", events=())
    cfg = {"alerts": {"recipients": ["a@x"]}, "smtp": {"host": "mail.x"}}
    # Must not raise — channel isolation.
    EmailChannel().send(cfg, object(), note)


def test_channels_registry_contains_email():
    from ruckus_dashboard.notify.channels import CHANNELS
    assert "email" in CHANNELS
    assert CHANNELS["email"].name == "email"
